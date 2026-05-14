"""Excel 导出器。

接受一组"已合并"的论文记录(每条 dict 满足下方 schema),
输出到 results/{slug}_{timestamp}.xlsx。

期望的记录 schema(由 search.py 合并三个 API 后产出):
  title, first_author, authors_full, year, journal,
  volume, issue, doctype,
  abstract, abstract_source,   # source ∈ {"Crossref", "OpenAlex", "无"}
  keywords,
  scopus_cited_by, crossref_cited_by, reference_count,
  doi, eid
"""

import os
import re
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
HEADER_FONT = Font(bold=True)
LINK_FONT = Font(color="0563C1", underline="single")

SOURCE_FILLS = {
    "Crossref": PatternFill(start_color="DFF0D8", end_color="DFF0D8", fill_type="solid"),
    "OpenAlex": PatternFill(start_color="E5D5F0", end_color="E5D5F0", fill_type="solid"),
    "无":       PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid"),
}

# (表头, 字段名 / 特殊键, 列宽)
COLUMNS: list[tuple[str, str | None, int]] = [
    ("序号",          None,                 6),
    ("标题",          "title",             60),
    ("第一作者",      "first_author",      18),
    ("全部作者",      "authors_full",      40),
    ("年份",          "year",               8),
    ("期刊",          "journal",           28),
    ("卷期",          "_vol_issue",        12),
    ("文献类型",      "doctype",           14),
    ("摘要",          "abstract",          80),
    ("摘要来源",      "abstract_source",   12),
    ("关键词",        "keywords",          30),
    ("Scopus 被引数",   "scopus_cited_by",   14),
    ("Crossref 被引数", "crossref_cited_by", 16),
    ("参考文献数",    "reference_count",   12),
    ("DOI",           "doi",               28),
    ("Scopus 链接",   "_scopus_url",       45),
    ("DOI 链接",      "_doi_url",          35),
]


def export(records: list[dict], first_keyword: str, output_dir: str = "results") -> str:
    """生成 xlsx,返回绝对路径。"""
    os.makedirs(output_dir, exist_ok=True)

    slug = _slugify(first_keyword)
    ts = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    path = os.path.abspath(os.path.join(output_dir, f"{slug}_{ts}.xlsx"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Papers"

    _write_header(ws)
    for i, rec in enumerate(records, 1):
        _write_row(ws, row_idx=i + 1, seq=i, rec=rec)

    ws.freeze_panes = "A2"
    wb.save(path)
    return path


def _write_header(ws) -> None:
    for col_idx, (header, _, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 22


def _write_row(ws, row_idx: int, seq: int, rec: dict) -> None:
    for col_idx, (_, field, _) in enumerate(COLUMNS, 1):
        if field is None:
            value = seq
        else:
            value = _resolve(rec, field)
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        _style_cell(cell, field, value)


def _style_cell(cell, field: str | None, value) -> None:
    if field == "abstract":
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        return
    if field == "abstract_source":
        fill = SOURCE_FILLS.get(value)
        if fill:
            cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        return
    if field in ("_scopus_url", "_doi_url"):
        if value:
            cell.hyperlink = value
            cell.font = LINK_FONT
        cell.alignment = Alignment(vertical="top")
        return
    cell.alignment = Alignment(vertical="top")


def _resolve(rec: dict, field: str):
    if field == "_vol_issue":
        vol = str(rec.get("volume") or "").strip()
        issue = str(rec.get("issue") or "").strip()
        if vol and issue:
            return f"{vol}({issue})"
        if vol:
            return vol
        if issue:
            return f"({issue})"
        return ""
    if field == "_scopus_url":
        eid = rec.get("eid") or ""
        return f"https://www.scopus.com/record/display.uri?eid={eid}" if eid else ""
    if field == "_doi_url":
        doi = rec.get("doi") or ""
        return f"https://doi.org/{doi}" if doi else ""
    if field in ("scopus_cited_by", "crossref_cited_by", "reference_count"):
        return _to_int_or_blank(rec.get(field))
    value = rec.get(field, "")
    return "" if value is None else value


def _to_int_or_blank(v):
    if v is None or v == "":
        return ""
    try:
        return int(v)
    except (TypeError, ValueError):
        return v


def _slugify(text: str) -> str:
    text = (text or "").strip()
    text = re.sub(r"\s+", "_", text)
    text = re.sub(r"[^\w\-]", "", text, flags=re.UNICODE)
    return text or "papers"
